import re, time, random, requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
]

def get_headers():
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'it-IT,it;q=0.9',
    }

def pausa():
    time.sleep(random.uniform(2, 4))

EMAIL_PATTERN = r'[a-zA-Z0-9.\-_+]+@[a-zA-Z0-9.\-_]+\.[a-z]{2,6}'
TEL_PATTERN = r'(?:(?:\+39|0039)?[\s\-]?)?(?:0\d{1,4}[\s\-]?\d{4,8}|\d{3}[\s\-]?\d{3}[\s\-]?\d{4})'

DOMINI_ESCLUDI = [
    'example.com','noreply','no-reply','privacy','webmaster',
    'sentry','w3.org','schema.org','googleapis','facebook',
    'twitter','youtube','google.','microsoft','apple.com',
    'wordpress','fontawesome','paginegialle','pec.it',
]

def is_email_valida(email):
    el = email.lower()
    if any(x in el for x in DOMINI_ESCLUDI): return False
    if len(email) < 8: return False
    if el.endswith(('.js','.css','.png','.jpg','.svg','.php')): return False
    return True

# Categorie aziende target per Toyota scaffalature/muletti
CATEGORIE = [
    ('magazzini-depositi', 'Magazzini e Depositi'),
    ('logistica', 'Logistica e Trasporti'),
    ('spedizioni', 'Spedizioni e Corrieri'),
    ('industria-manifatturiera', 'Industria Manifatturiera'),
    ('supermercati', 'Supermercati e GDO'),
    ('materiali-edili', 'Materiali Edili'),
    ('aziende-alimentari', 'Industria Alimentare'),
    ('ferramenta-ingrosso', 'Ferramenta Ingrosso'),
    ('ricambi-auto', 'Ricambi Auto Ingrosso'),
    ('farmaceutica', 'Farmaceutica e Parafarmacia'),
    ('carta-cartone', 'Carta e Cartone'),
    ('metalmeccanica', 'Metalmeccanica'),
    ('cooperative', 'Cooperative'),
    ('distribuzione-ingrosso', 'Distribuzione Ingrosso'),
    ('arredamento-ingrosso', 'Arredamento Ingrosso'),
]

PROVINCE = [
    ('pescara', 'Pescara'),
    ('teramo', 'Teramo'),
]

aziende = []
email_set = set()

print('='*60)
print('RICERCA AZIENDE TOYOTA - PESCARA E TERAMO')
print(f'Inizio: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
print('='*60)

for cat_url, cat_nome in CATEGORIE:
    for prov_url, prov_nome in PROVINCE:
        for pagina in range(1, 6):
            if pagina == 1:
                url = f'https://www.paginegialle.it/ricerca/{cat_url}/{prov_url}'
            else:
                url = f'https://www.paginegialle.it/ricerca/{cat_url}/{prov_url}/p-{pagina}'

            print(f'  [{prov_nome}] {cat_nome} - p.{pagina}')

            try:
                res = requests.get(url, headers=get_headers(), timeout=20)
                testo = res.text

                if res.status_code != 200 or 'nessun risultato' in testo.lower():
                    print(f'    Nessun risultato')
                    break

                # Estrai nomi aziende
                nomi = re.findall(r'class="[^"]*denomination[^"]*"[^>]*>([^<]+)<', testo)
                nomi += re.findall(r'itemprop="name"[^>]*>([^<]+)<', testo)

                # Estrai email
                emails = re.findall(EMAIL_PATTERN, testo)

                # Estrai telefoni
                telefoni = re.findall(TEL_PATTERN, testo)

                # Estrai indirizzi
                indirizzi = re.findall(r'itemprop="streetAddress"[^>]*>([^<]+)<', testo)

                # Associa dati
                nuovi = 0
                for j, email in enumerate(emails):
                    email = email.lower().strip('.')
                    if not is_email_valida(email): continue
                    if email in email_set: continue
                    email_set.add(email)

                    nome = nomi[j] if j < len(nomi) else f'Azienda {prov_nome}'
                    tel = telefoni[j].strip() if j < len(telefoni) else ''
                    indirizzo = indirizzi[j].strip() if j < len(indirizzi) else ''

                    aziende.append({
                        'nome': nome.strip(),
                        'email': email,
                        'telefono': tel,
                        'indirizzo': indirizzo,
                        'provincia': prov_nome,
                        'categoria': cat_nome,
                    })
                    nuovi += 1
                    print(f'    ✓ {nome.strip()} - {email}')

                # Anche senza email salva le aziende con solo telefono
                for j, nome in enumerate(nomi):
                    nome = nome.strip()
                    if not nome or len(nome) < 4: continue
                    tel = telefoni[j].strip() if j < len(telefoni) else ''
                    indirizzo = indirizzi[j].strip() if j < len(indirizzi) else ''

                    # Controlla se già presente
                    già_presente = any(a['nome'] == nome for a in aziende)
                    if già_presente: continue

                    aziende.append({
                        'nome': nome,
                        'email': '',
                        'telefono': tel,
                        'indirizzo': indirizzo,
                        'provincia': prov_nome,
                        'categoria': cat_nome,
                    })

                print(f'    +{nuovi} email | TOTALE: {len(aziende)}')
                if nuovi == 0 and pagina >= 2: break

            except Exception as e:
                print(f'    Errore: {e}')
                break

            pausa()

print(f'\nTOTALE AZIENDE: {len(aziende)}')
con_email = sum(1 for a in aziende if a['email'])
print(f'Con email: {con_email}')
print(f'Solo telefono: {len(aziende) - con_email}')

# Crea Excel
wb = Workbook()
ws = wb.active
ws.title = "Aziende Toyota"

header_fill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
green_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
gray_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
yellow_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
pe_fill = PatternFill(start_color='e8f4fd', end_color='e8f4fd', fill_type='solid')
te_fill = PatternFill(start_color='fef9e7', end_color='fef9e7', fill_type='solid')

intestazioni = ['#', 'Nome Azienda', 'Email', 'Telefono', 'Indirizzo', 'Provincia', 'Categoria', 'Stato', 'Note']

for col, titolo in enumerate(intestazioni, 1):
    cella = ws.cell(row=1, column=col, value=titolo)
    cella.font = header_font
    cella.fill = header_fill
    cella.alignment = Alignment(horizontal='center', vertical='center')

ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

for i, a in enumerate(aziende, 1):
    row = i + 1
    ha_email = bool(a['email'])
    
    if ha_email:
        fill = green_fill
    elif a['provincia'] == 'Pescara':
        fill = pe_fill
    else:
        fill = te_fill

    ws.cell(row=row, column=1, value=i).fill = fill
    ws.cell(row=row, column=2, value=a['nome']).fill = fill

    if ha_email:
        ec = ws.cell(row=row, column=3, value=a['email'])
        ec.fill = fill
        ec.font = Font(color='0563C1', underline='single')
    else:
        ws.cell(row=row, column=3, value='').fill = fill

    ws.cell(row=row, column=4, value=a['telefono']).fill = fill
    ws.cell(row=row, column=5, value=a['indirizzo']).fill = fill
    ws.cell(row=row, column=6, value=a['provincia']).fill = fill
    ws.cell(row=row, column=7, value=a['categoria']).fill = fill

    stato = ws.cell(row=row, column=8, value='Da contattare')
    stato.fill = yellow_fill
    stato.font = Font(bold=True, color='856404')

    ws.cell(row=row, column=9, value='').fill = fill

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 16
ws.column_dimensions['I'].width = 30

nome_file = f'aziende_toyota_pescara_teramo_{datetime.now().strftime("%d%m%Y")}.xlsx'
wb.save(nome_file)
print(f'\nFile salvato: {nome_file}')
