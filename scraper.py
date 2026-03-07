import subprocess, sys
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', 'pdfplumber', 'requests'])

import os, requests, re, time, random, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

SPECIALIZZAZIONI = {
    'allergologia':         'Allergologia ed Immunologia Clinica',
    'dermatologia':         'Dermatologia e Venerologia',
    'gastroenterologia':    'Gastroenterologia',
    'respiratorio':         "Malattie dell'Apparato Respiratorio",
    'medicina interna':     'Medicina Interna',
    'pediatria':            'Pediatria',
    'reumatologia':         'Reumatologia',
    'cardiochirurgia':      'Cardiochirurgia',
    'chirurgia generale':   'Chirurgia Generale',
    'otorinolaringoiatria': 'Otorinolaringoiatria',
    'anatomia patologica':  'Anatomia Patologica',
    'medicina del lavoro':  'Medicina del Lavoro e Sicurezza',
    'medico di famiglia':   'Medicina Generale (Medici di Famiglia)',
    'direzione medica':     'Direzione Medica di Presidio Ospedaliero',
    'medicina comunita':    'Medicina di Comunita e Cure Primarie',
    'farmacista':           'Farmacista',
}

SPEC_KEYWORDS = {
    'allergologia':         ['allergolog', 'immunolog', 'allergia'],
    'dermatologia':         ['dermatolog', 'venerolog'],
    'gastroenterologia':    ['gastroenterolog', 'gastroentero'],
    'respiratorio':         ['respirator', 'pneumolog', 'polmonare'],
    'medicina interna':     ['medicina interna', 'internista'],
    'pediatria':            ['pediatr'],
    'reumatologia':         ['reumatolog'],
    'cardiochirurgia':      ['cardiochirurg'],
    'chirurgia generale':   ['chirurgia generale', 'chirurgo generale'],
    'otorinolaringoiatria': ['otorinolaringoiatr', 'orl', 'otorino'],
    'anatomia patologica':  ['anatomia patologica', 'patologo'],
    'medicina del lavoro':  ['medicina del lavoro', 'medico del lavoro'],
    'medico di famiglia':   ['medico di famiglia', 'medicina generale', 'mmg'],
    'direzione medica':     ['direttore medico', 'direzione medica'],
    'medicina comunita':    ['cure primarie', 'medicina territorio'],
    'farmacista':           ['farmacist', 'farmacia'],
}

PROVINCE = {
    'pescara': 'PESCARA',
    'chieti': 'CHIETI',
    'teramo': 'TERAMO',
    "l'aquila": "L'AQUILA",
    'lanciano': 'CHIETI',
    'vasto': 'CHIETI',
    'avezzano': "L'AQUILA",
    'sulmona': "L'AQUILA",
}

USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
]

def get_headers():
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'it-IT,it;q=0.9',
        'Referer': 'https://www.google.it/',
    }

def pausa(min=1.5, max=3.5):
    time.sleep(random.uniform(min, max))

EMAIL_PATTERN = r'[a-zA-Z0-9.\-_+]+@[a-zA-Z0-9.\-_]+\.[a-z]{2,6}'

DOMINI_ESCLUDI = [
    'example.com','noreply','no-reply','privacy@','dpo@','webmaster',
    'sentry.io','w3.org','schema.org','googleapis','jquery','bootstrap',
    'cloudflare','facebook','twitter','youtube','google.com','microsoft',
    'apple.com','wix.com','wordpress','jsdelivr','cdnjs','fontawesome',
    'sitemaps','robots','sitemap','test@','admin@',
]

def is_email_valida(email):
    el = email.lower()
    if any(x in el for x in DOMINI_ESCLUDI): return False
    if len(email) < 8: return False
    if el.endswith(('.js','.css','.png','.jpg','.svg','.php')): return False
    if el.count('@') != 1: return False
    return True

def rileva_specializzazione(testo):
    tl = testo.lower()
    for chiave, kws in SPEC_KEYWORDS.items():
        if any(kw in tl for kw in kws):
            return chiave
    return None

def estrai_email_con_spec(testo, provincia, spec_forzata=None):
    trovate = []
    for match in re.finditer(EMAIL_PATTERN, testo):
        email = match.group().lower().strip('.')
        if not is_email_valida(email): continue
        if spec_forzata:
            spec = spec_forzata
        else:
            contesto = testo[max(0,match.start()-500):match.end()+500]
            spec = rileva_specializzazione(contesto)
        if spec:
            nome = email.split('@')[0].replace('.',' ').replace('_',' ').replace('-',' ').title()
            r = [datetime.now().strftime('%d/%m/%Y'), nome, email,
                 SPECIALIZZAZIONI[spec], provincia.upper()]
            if r not in trovate: trovate.append(r)
    return trovate

def crea_browser():
    if not SELENIUM_OK: return None
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument(f'--user-agent={random.choice(USER_AGENTS)}')
    options.add_argument('--window-size=1920,1080')
    try:
        driver = webdriver.Chrome(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        return driver
    except Exception as e:
        print(f'Browser non disponibile: {e}')
        return None

def scarica(url, driver=None, usa_selenium=False):
    if usa_selenium and driver:
        try:
            driver.get(url)
            time.sleep(3)
            return driver.page_source
        except: return ''
    try:
        res = requests.Session().get(url, headers=get_headers(), timeout=25)
        if res.status_code in (403, 429, 503) and driver:
            driver.get(url)
            time.sleep(3)
            return driver.page_source
        if res.status_code == 200:
            return res.text
        return ''
    except Exception as e:
        print(f'    Errore scarica: {e}')
        return ''

def scarica_pdf(url):
    if not PDF_OK: return ''
    try:
        res = requests.get(url, headers=get_headers(), timeout=30)
        if res.status_code != 200: return ''
        with pdfplumber.open(io.BytesIO(res.content)) as pdf:
            testo = ''
            for pagina in pdf.pages:
                t = pagina.extract_text()
                if t: testo += t + '\n'
        return testo
    except Exception as e:
        print(f'    Errore PDF: {e}')
        return ''

# ============================================================
# FONTE 1: INI-PEC (Registro pubblico PEC - fonte ufficiale)
# ============================================================
def scrapa_inipec(driver=None):
    trovate = []
    print('\n  Ricerca su INI-PEC...')
    categorie = [
        ('medico chirurgo', 'medico di famiglia'),
        ('farmacista', 'farmacista'),
        ('medico specialista', 'medicina interna'),
    ]
    province_codici = ['PE', 'CH', 'TE', 'AQ']
    for cat, spec_chiave in categorie:
        for prov in province_codici:
            url = f'https://www.inipec.gov.it/cerca-pec/-/pec/professionisti?categoria={cat.replace(" ","+")}&provincia={prov}'
            print(f'    INI-PEC: {cat} - {prov}')
            t = scarica(url, driver=driver, usa_selenium=True)
            if not t: pausa(); continue
            for r in estrai_email_con_spec(t, prov, spec_chiave):
                if r not in trovate: trovate.append(r)
            pausa()
    return trovate

# ============================================================
# FONTE 2: FNOMCeO - Albo medici ufficiale
# ============================================================
def scrapa_fnomceo(driver=None):
    trovate = []
    print('\n  Ricerca su FNOMCeO...')
    province = ['Pescara', 'Chieti', 'Teramo', "L'Aquila"]
    for prov in province:
        url = f'https://portale.fnomceo.it/trova-medico/?provincia={prov}'
        print(f'    FNOMCeO: {prov}')
        t = scarica(url, driver=driver, usa_selenium=True)
        if not t: pausa(); continue
        for r in estrai_email_con_spec(t, prov):
            if r not in trovate: trovate.append(r)
        pausa()
    return trovate

# ============================================================
# FONTE 3: Dottori.it
# ============================================================
def scrapa_dottori(spec_chiave, prov, driver=None):
    trovate = []
    termine = SPEC_KEYWORDS[spec_chiave][0].replace(' ', '-')
    p = prov.replace("'", '-').replace(' ', '-').lower()
    for pagina in range(1, 6):
        url = f'https://www.dottori.it/{termine}/{p}/?page={pagina}'
        print(f'    Dottori.it: {SPECIALIZZAZIONI[spec_chiave]} - {prov} p.{pagina}')
        t = scarica(url, driver=driver, usa_selenium=True)
        if not t: break
        agg = 0
        for email in set(re.findall(EMAIL_PATTERN, t)):
            email = email.lower().strip('.')
            if is_email_valida(email):
                nome = email.split('@')[0].replace('.', ' ').replace('_', ' ').title()
                r = [datetime.now().strftime('%d/%m/%Y'), nome, email,
                     SPECIALIZZAZIONI[spec_chiave], prov.upper()]
                if r not in trovate: trovate.append(r); agg += 1
        print(f'      +{agg} email')
        if agg == 0 and pagina > 2: break
        pausa()
    return trovate

# ============================================================
# FONTE 4: MioDottore
# ============================================================
def scrapa_miodottore(spec_chiave, prov, driver=None):
    if not driver: return []
    trovate = []
    termine = SPEC_KEYWORDS[spec_chiave][0].replace(' ', '-')
    p = prov.replace("'", '-').replace(' ', '-').lower()
    for pagina in range(1, 6):
        url = f'https://www.miodottore.it/{termine}/{p}?page={pagina}'
        print(f'    MioDottore: {SPECIALIZZAZIONI[spec_chiave]} - {prov} p.{pagina}')
        t = scarica(url, driver=driver, usa_selenium=True)
        if not t: break
        agg = 0
        for email in set(re.findall(EMAIL_PATTERN, t)):
            email = email.lower().strip('.')
            if is_email_valida(email):
                nome = email.split('@')[0].replace('.', ' ').replace('_', ' ').title()
                r = [datetime.now().strftime('%d/%m/%Y'), nome, email,
                     SPECIALIZZAZIONI[spec_chiave], prov.upper()]
                if r not in trovate: trovate.append(r); agg += 1
        print(f'      +{agg} email')
        if agg == 0 and pagina > 2: break
        pausa()
    return trovate

# ============================================================
# FONTE 5: Google Maps (studi medici con email pubblica)
# ============================================================
def scrapa_google_maps(spec_chiave, prov, driver=None):
    if not driver: return []
    trovate = []
    termine = SPEC_KEYWORDS[spec_chiave][0]
    query = f'{termine} {prov} abruzzo'
    url = f'https://www.google.com/maps/search/{query.replace(" ", "+")}'
    print(f'    Google Maps: {termine} - {prov}')
    try:
        driver.get(url)
        time.sleep(4)
        t = driver.page_source
        for r in estrai_email_con_spec(t, prov, spec_chiave):
            if r not in trovate: trovate.append(r)
    except Exception as e:
        print(f'      Errore Maps: {e}')
    pausa()
    return trovate

# ============================================================
# FONTE 6: Siti ordini provinciali e ASL
# ============================================================
SITI_UFFICIALI = [
    ('https://www.omceope.it', 'PESCARA'),
    ('https://www.omceochieti.it', 'CHIETI'),
    ('https://www.omceoteramo.it', 'TERAMO'),
    ('https://www.omceoaq.it', "L'AQUILA"),
    ('https://www.aslpe.it', 'PESCARA'),
    ('https://www.asl2abruzzo.it', 'CHIETI'),
    ('https://www.asl1abruzzo.it', "L'AQUILA"),
    ('https://www.aslteramo.it', 'TERAMO'),
    ('https://www.sanita.regione.abruzzo.it', 'ABRUZZO'),
    ('https://www.farmacistiabruzzo.it', 'ABRUZZO'),
    ('https://www.ordinefarmacistipescara.it', 'PESCARA'),
    ('https://www.ordinefarmacistichieti.it', 'CHIETI'),
]

PDF_UFFICIALI = [
    ('https://www.omceope.it/albo/elenco_iscritti.pdf', 'PESCARA'),
    ('https://www.omceochieti.it/albo/iscritti.pdf', 'CHIETI'),
    ('https://www.omceoteramo.it/albo/elenco.pdf', 'TERAMO'),
    ('https://www.omceoaq.it/albo/elenco_iscritti.pdf', "L'AQUILA"),
    ('https://www.aslpe.it/documenti/medici_mmg.pdf', 'PESCARA'),
    ('https://www.farmacistiabruzzo.it/albo/iscritti.pdf', 'ABRUZZO'),
]

def scrapa_sito_ufficiale(url, provincia, driver=None):
    trovate, visitati, da_visitare = [], set(), [url]
    base = '/'.join(url.split('/')[:3])
    KW = ['medic','dott','staff','specialist','contatt','reparti','ambulatori',
          'personale','farmac','chirurg','cardio','gastro','dermato','pneumo',
          'reumato','allergo','pediatr','direttore','primario','pdf','elenco','albo']
    livello = 0
    while da_visitare and livello <= 2:
        batch = da_visitare[:]; da_visitare = []
        for pu in batch:
            if pu in visitati or len(visitati) > 50: continue
            visitati.add(pu)
            if pu.lower().endswith('.pdf'):
                testo = scarica_pdf(pu)
                for r in estrai_email_con_spec(testo, provincia):
                    if r not in trovate: trovate.append(r)
                pausa(); continue
            t = scarica(pu, driver=driver)
            if not t: pausa(); continue
            for r in estrai_email_con_spec(t, provincia):
                if r not in trovate: trovate.append(r)
            if livello < 2:
                for lk in re.findall(r'href=["\']([^"\']+)["\']', t):
                    if any(k in lk.lower() for k in KW):
                        fl = (lk if lk.startswith('http')
                              else base+lk if lk.startswith('/')
                              else base+'/'+lk)
                        if base in fl and fl not in visitati:
                            da_visitare.append(fl)
            pausa()
        livello += 1
    return trovate

# ============================================================
# MAIN
# ============================================================
print('=' * 60)
print('RICERCA EMAIL SANITARI - ABRUZZO v2.0')
print(f'Inizio: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
print('=' * 60)

dati_finali = []
email_set = set()

def aggiungi(records):
    n = 0
    for r in records:
        if r[2] not in email_set:
            email_set.add(r[2])
            dati_finali.append(r)
            n += 1
    if n > 0:
        print(f'    >>> +{n} nuove | TOTALE: {len(dati_finali)}')
    return n

print('\nAvvio browser...')
driver = crea_browser()
print('  Browser OK' if driver else '  Uso solo requests')

province_principali = ['pescara', 'chieti', 'teramo', "l'aquila"]

try:
    print('\n[FASE 1] PDF ufficiali...')
    for url, prov in PDF_UFFICIALI:
        print(f'  {url}')
        testo = scarica_pdf(url)
        aggiungi(estrai_email_con_spec(testo, prov))

    print('\n[FASE 2] Siti ordini e ASL...')
    for url, prov in SITI_UFFICIALI:
        print(f'  {url}')
        aggiungi(scrapa_sito_ufficiale(url, prov, driver=driver))

    print('\n[FASE 3] INI-PEC (registro PEC ufficiale)...')
    aggiungi(scrapa_inipec(driver=driver))

    print('\n[FASE 4] FNOMCeO (albo medici nazionale)...')
    aggiungi(scrapa_fnomceo(driver=driver))

    print('\n[FASE 5] Dottori.it...')
    for spec in SPECIALIZZAZIONI:
        for prov in province_principali:
            aggiungi(scrapa_dottori(spec, prov, driver=driver))

    print('\n[FASE 6] MioDottore...')
    for spec in SPECIALIZZAZIONI:
        for prov in province_principali:
            aggiungi(scrapa_miodottore(spec, prov, driver=driver))

    print('\n[FASE 7] Google Maps...')
    for spec in list(SPECIALIZZAZIONI.keys())[:8]:
        for prov in province_principali:
            aggiungi(scrapa_google_maps(spec, prov, driver=driver))

finally:
    if driver:
        driver.quit()

# Salva Excel
print(f'\n{"="*60}')
print(f'TOTALE FINALE: {len(dati_finali)} email')
print('\nRiepilogo per specializzazione:')
for label in SPECIALIZZAZIONI.values():
    c = sum(1 for r in dati_finali if r[3] == label)
    if c > 0: print(f'  {label}: {c}')

nome_file = f"email_sanitari_abruzzo_{datetime.now().strftime('%d%m%Y')}.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "Email Sanitari Abruzzo"

# Intestazioni con stile
intestazioni = ['Data', 'Nome', 'Email', 'Specializzazione', 'Provincia']
for col, titolo in enumerate(intestazioni, 1):
    cella = ws.cell(row=1, column=col, value=titolo)
    cella.font = Font(bold=True, color='FFFFFF')
    cella.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    cella.alignment = Alignment(horizontal='center')

for riga in dati_finali:
    ws.append(riga)

# Larghezza colonne
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 15

wb.save(nome_file)
print(f'\nFile Excel salvato: {nome_file}')
print(f'Scaricalo dalla sezione Artifacts del workflow!')
