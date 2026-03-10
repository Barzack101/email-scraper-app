import os, re, time, random
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    SELENIUM_OK = True
except:
    SELENIUM_OK = False

USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36',
]

def get_headers():
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'it-IT,it;q=0.9',
        'Referer': 'https://www.google.it/',
    }

def pausa():
    time.sleep(random.uniform(2, 4))

EMAIL_PATTERN = r'[a-zA-Z0-9.\-_+]+@[a-zA-Z0-9.\-_]+\.[a-z]{2,6}'

DOMINI_ESCLUDI = [
    'example.com','noreply','no-reply','privacy','webmaster',
    'sentry','w3.org','schema.org','googleapis','facebook',
    'twitter','youtube','google.','microsoft','apple.com',
    'wordpress','fontawesome','gstatic','paginegialle',
    'pagineinfoimprese','tuttitalia',
]

def is_email_valida(email):
    el = email.lower()
    if any(x in el for x in DOMINI_ESCLUDI): return False
    if len(email) < 8: return False
    if el.endswith(('.js','.css','.png','.jpg','.svg','.php')): return False
    return True

def crea_browser():
    if not SELENIUM_OK: return None
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_argument(f'--user-agent={random.choice(USER_AGENTS)}')
    options.add_argument('--window-size=1920,1080')
    try:
        driver = webdriver.Chrome(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        return driver
    except Exception as e:
        print(f'Browser non disponibile: {e}')
        return None

def scarica(url, driver=None):
    if driver:
        try:
            driver.get(url)
            time.sleep(3)
            return driver.page_source
        except: return ''
    try:
        res = requests.get(url, headers=get_headers(), timeout=20)
        return res.text if res.status_code == 200 else ''
    except: return ''

def estrai_info_agenzia(testo, citta, tipo):
    risultati = []
    # Cerca email
    for match in re.finditer(EMAIL_PATTERN, testo):
        email = match.group().lower().strip('.')
        if not is_email_valida(email): continue
        # Contesto attorno all'email
        contesto = testo[max(0,match.start()-400):match.end()+400]
        contesto_pulito = re.sub(r'<[^>]+>', ' ', contesto)
        contesto_pulito = re.sub(r'\s+', ' ', contesto_pulito).strip()
        # Cerca nome agenzia nel contesto
        nome = contesto_pulito[:80].strip()
        risultati.append({
            'nome': nome,
            'email': email,
            'tipo': tipo,
            'citta': citta,
        })
    return risultati

# Fonti da cercare
CITTA = [
    'Milano', 'Roma', 'Napoli', 'Torino', 'Bologna', 'Firenze',
    'Venezia', 'Genova', 'Palermo', 'Bari', 'Pescara', 'Chieti',
    'Teramo', 'Ancona', 'Perugia', 'Verona', 'Padova', 'Brescia',
    'Modena', 'Parma', 'Catania', 'Cagliari', 'Trento', 'Trieste'
]

TIPI_AGENZIA = [
    'agenzia comunicazione',
    'agenzia marketing',
    'agenzia eventi',
    'agenzia pubblicitaria',
    'studio comunicazione',
    'web agency',
]

agenzie = []
email_set = set()

print('='*60)
print('RICERCA AGENZIE MARKETING E COMUNICAZIONE - ITALIA')
print(f'Inizio: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
print('='*60)

print('\nAvvio browser...')
driver = crea_browser()
print('Browser OK' if driver else 'Solo requests')

try:
    # FONTE 1: Pagine Gialle
    print('\n[FONTE 1] Pagine Gialle...')
    for tipo in TIPI_AGENZIA:
        for citta in CITTA:
            termine = tipo.replace(' ', '-')
            citta_url = citta.lower().replace(' ', '-').replace("'", '')
            for pagina in range(1, 4):
                if pagina == 1:
                    url = f'https://www.paginegialle.it/ricerca/{termine}/{citta_url}'
                else:
                    url = f'https://www.paginegialle.it/ricerca/{termine}/{citta_url}/p-{pagina}'
                
                print(f'  PG: {tipo} - {citta} p.{pagina}')
                testo = scarica(url, driver=driver)
                if not testo: pausa(); break
                
                nuovi = 0
                for r in estrai_info_agenzia(testo, citta, tipo):
                    if r['email'] not in email_set:
                        email_set.add(r['email'])
                        agenzie.append(r)
                        nuovi += 1
                
                print(f'    +{nuovi} | TOTALE: {len(agenzie)}')
                if nuovi == 0 and pagina > 1: break
                pausa()

    # FONTE 2: Siti diretti delle agenzie trovate su Google
    print('\n[FONTE 2] Ricerca Google siti agenzie...')
    for tipo in TIPI_AGENZIA[:3]:  # Solo primi 3 tipi per velocità
        for citta in CITTA[:12]:   # Solo prime 12 città
            query = f'"{tipo}" "{citta}" site:it email'
            url = f'https://www.google.com/search?q={query.replace(" ", "+")}'
            print(f'  Google: {tipo} - {citta}')
            testo = scarica(url, driver=driver)
            if not testo: pausa(); continue
            
            # Estrai URL dei siti trovati
            siti = re.findall(r'https?://(?!www\.google)[a-zA-Z0-9.\-]+\.[a-z]{2,6}(?:/[^\s"<>]*)?', testo)
            siti = list(set([s.split('?')[0] for s in siti if len(s) < 100]))[:5]
            
            for sito in siti:
                print(f'    Visito: {sito}')
                testo_sito = scarica(sito, driver=driver)
                if not testo_sito: continue
                for r in estrai_info_agenzia(testo_sito, citta, tipo):
                    if r['email'] not in email_set:
                        email_set.add(r['email'])
                        agenzie.append(r)
                        print(f'      ✓ {r["email"]}')
                pausa()
            pausa()

    # FONTE 3: Elenco imprese
    print('\n[FONTE 3] Elenco imprese...')
    for tipo in TIPI_AGENZIA[:2]:
        for citta in CITTA[:10]:
            query = tipo.replace(' ', '+')
            citta_q = citta.replace(' ', '+')
            url = f'https://www.elencaimprese.it/ricerca?q={query}&where={citta_q}'
            print(f'  EI: {tipo} - {citta}')
            testo = scarica(url, driver=driver)
            if testo:
                for r in estrai_info_agenzia(testo, citta, tipo):
                    if r['email'] not in email_set:
                        email_set.add(r['email'])
                        agenzie.append(r)
            pausa()

finally:
    if driver: driver.quit()

print(f'\nTOTALE AGENZIE TROVATE: {len(agenzie)}')

# Crea Excel professionale
wb = Workbook()
ws = wb.active
ws.title = "Agenzie da Contattare"

header_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
green_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
yellow_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
blue_fill = PatternFill(start_color='cce5ff', end_color='cce5ff', fill_type='solid')

intestazioni = ['#', 'Email', 'Tipo Agenzia', 'Città', 'Stato Contatto', 'LinkedIn', 'Note']

for col, titolo in enumerate(intestazioni, 1):
    cella = ws.cell(row=1, column=col, value=titolo)
    cella.font = header_font
    cella.fill = header_fill
    cella.alignment = Alignment(horizontal='center')

ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

for i, a in enumerate(agenzie, 1):
    row = i + 1
    tipo = a['tipo'].title()
    citta = a['citta']
    
    ws.cell(row=row, column=1, value=i).fill = green_fill
    ws.cell(row=row, column=2, value=a['email']).fill = green_fill
    ws.cell(row=row, column=3, value=tipo).fill = green_fill
    ws.cell(row=row, column=4, value=citta).fill = green_fill
    
    stato = ws.cell(row=row, column=5, value='Da contattare')
    stato.fill = yellow_fill
    stato.font = Font(bold=True, color='856404')
    
    linkedin_url = f'https://www.linkedin.com/search/results/people/?keywords={tipo.replace(" ", "%20")}%20{citta.replace(" ", "%20")}'
    li = ws.cell(row=row, column=6, value='🔗 Cerca su LinkedIn')
    li.hyperlink = linkedin_url
    li.font = Font(color='0A66C2', underline='single')
    li.fill = blue_fill
    
    ws.cell(row=row, column=7, value='').fill = green_fill

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 30

# Secondo foglio: messaggi pronti
ws2 = wb.create_sheet("Messaggi LinkedIn")
ws2['A1'] = 'MESSAGGIO 1 - PRIMO CONTATTO'
ws2['A1'].font = Font(bold=True, size=12, color='1a3a5c')
ws2['A2'] = '''Gentile [Nome],

Mi permetto di contattarla in quanto la sua agenzia opera nel settore della comunicazione e del marketing, ambito in cui posso offrire un supporto concreto.

Sono Carlo D'Attanasio, fondatore di CD Artificial Innovation, specializzata nella fornitura di database di contatti professionali verificati per qualsiasi settore — dal medicale all'ingegneria, dalla tecnologia alla ristorazione.

Forniamo liste Excel aggiornate con nome, email, settore e provincia, provenienti da fonti ufficiali pubbliche. Dati pronti per campagne marketing, eventi e attività commerciali.

Se ritiene possa essere di interesse, sarei lieto di inviarle un campione gratuito di 50 contatti nel settore di suo interesse, senza alcun impegno.

Cordiali saluti,
Carlo D'Attanasio
CD Artificial Innovation
https://barzack101.github.io/email-scraper-app/'''
ws2['A2'].alignment = Alignment(wrap_text=True)
ws2.row_dimensions[2].height = 200
ws2.column_dimensions['A'].width = 80

ws2['A4'] = 'MESSAGGIO 2 - FOLLOW UP (dopo 7 giorni)'
ws2['A4'].font = Font(bold=True, size=12, color='1a3a5c')
ws2['A5'] = '''Gentile [Nome],

Mi permetto di riprendere contatto riguardo al mio precedente messaggio.

Volevo ricordarle che l'offerta per il campione gratuito di 50 contatti verificati è ancora disponibile — un'opportunità per valutare concretamente la qualità del nostro servizio senza alcun impegno.

Rimango a disposizione per qualsiasi informazione.

Cordiali saluti,
Carlo D'Attanasio
CD Artificial Innovation'''
ws2['A5'].alignment = Alignment(wrap_text=True)
ws2.row_dimensions[5].height = 150

nome_file = f'agenzie_da_contattare_{datetime.now().strftime("%d%m%Y")}.xlsx'
wb.save(nome_file)
print(f'File salvato: {nome_file}')
print(f'Foglio 1: {len(agenzie)} agenzie')
print('Foglio 2: Messaggi LinkedIn pronti')
